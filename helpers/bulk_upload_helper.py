import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date
from io import BytesIO
from django.http import HttpResponse
from django.db import transaction as db_transaction
from finance.models import Transaction

from serializers.base_serializers import TransactionCreateSerializer


# Display name -> DB field name mapping
COLUMN_MAPPING = {
    'TransactionTitle': 'title',
    'Transaction Amount': 'amount',
    'Type': 'transaction_type',
    'Date': 'date',
    'Category': 'category',
}

DISPLAY_COLUMNS = list(COLUMN_MAPPING.keys())
DB_FIELDS = list(COLUMN_MAPPING.values())

TRANSACTION_TYPES = ['Income', 'Expense']

CATEGORIES = [
    'rental', 'food/grocery', 'utilities/bills', 'entertainment', 'EMIs',
    'education', 'health', 'travel', 'personal expenses', 'other',
    'investment/SIPs', 'salary', 'incentives/bonus',
]

MAX_DATA_ROWS = 500


def generate_transaction_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # Header row — bold only
    header_font = openpyxl.styles.Font(bold=True)
    for col_idx, col_name in enumerate(DISPLAY_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    # Dropdown for Type column (column 3)
    type_col_letter = openpyxl.utils.get_column_letter(DISPLAY_COLUMNS.index('Type') + 1)
    type_formula = '"' + ','.join(TRANSACTION_TYPES) + '"'
    type_validation = DataValidation(type='list', formula1=type_formula, allow_blank=False)
    type_validation.error = 'Please select either Income or Expense.'
    type_validation.errorTitle = 'Invalid Type'
    type_validation.prompt = 'Select transaction type'
    type_validation.promptTitle = 'Type'
    ws.add_data_validation(type_validation)
    type_validation.add(f'{type_col_letter}2:{type_col_letter}{MAX_DATA_ROWS + 1}')

    # Dropdown for Category column (column 5)
    category_col_letter = openpyxl.utils.get_column_letter(DISPLAY_COLUMNS.index('Category') + 1)
    category_formula = '"' + ','.join(CATEGORIES) + '"'
    category_validation = DataValidation(type='list', formula1=category_formula, allow_blank=False)
    category_validation.error = 'Please select a valid category.'
    category_validation.errorTitle = 'Invalid Category'
    category_validation.prompt = 'Select category'
    category_validation.promptTitle = 'Category'
    ws.add_data_validation(category_validation)
    category_validation.add(f'{category_col_letter}2:{category_col_letter}{MAX_DATA_ROWS + 1}')

    # Format Date column as date so Excel shows calendar picker
    date_col_idx = DISPLAY_COLUMNS.index('Date') + 1
    date_col_letter = openpyxl.utils.get_column_letter(date_col_idx)
    for row_num in range(2, MAX_DATA_ROWS + 2):
        cell = ws.cell(row=row_num, column=date_col_idx)
        cell.number_format = 'YYYY-MM-DD'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transaction_template.xlsx"'
    return response


def normalize_date(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')

    # Try parsing common string formats
    value_str = str(value).strip()
    date_formats = [
        '%Y-%m-%d',       # 2026-01-15
        '%d-%m-%Y',       # 15-01-2026
        '%m/%d/%Y',       # 01/15/2026
        '%d/%m/%Y',       # 15/01/2026
        '%Y/%m/%d',       # 2026/01/15
    ]
    for fmt in date_formats:
        try:
            return datetime.strptime(value_str, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue

    # Return as-is and let the serializer raise the validation error
    return value_str


def process_bulk_upload(file, user):
    try:
        wb = openpyxl.load_workbook(file)
    except Exception:
        return False, {'error': 'Invalid file. Please upload a valid .xlsx file.'}

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header

    # Filter out completely empty rows
    data_rows = []
    for row in rows:
        if all(cell is None or str(cell).strip() == '' for cell in row):
            continue
        data_rows.append(row)

    if not data_rows:
        return False, {'error': 'No data rows found in the uploaded file.'}

    errors = {}
    validated_entries = []

    for row_idx, row in enumerate(data_rows, start=1):
        # Pad with None if fewer columns than expected
        if len(row) < len(DB_FIELDS):
            row = list(row) + [None] * (len(DB_FIELDS) - len(row))

        # Map display columns to DB field names
        row_data = {}
        for col_idx, db_field in enumerate(DB_FIELDS):
            value = row[col_idx]
            if db_field == 'date':
                row_data[db_field] = normalize_date(value)
            elif value is not None:
                row_data[db_field] = str(value).strip()
            else:
                row_data[db_field] = None

        serializer = TransactionCreateSerializer(data=row_data, context={'user': user})
        if serializer.is_valid():
            validated_entries.append(serializer.validated_data)
        else:
            errors[f'row_{row_idx}'] = serializer.errors

    # If any row had errors, return all errors and save nothing
    if errors:
        return False, {
            'error': 'Validation failed. No transactions were saved.',
            'row_errors': errors,
            'total_rows': len(data_rows),
            'failed_rows': len(errors),
        }

    # All rows valid — save atomically
    try:
        with db_transaction.atomic():
            created = []
            for entry in validated_entries:
                txn = Transaction.objects.create(user=user, **entry)
                created.append({
                    'id': txn.id,
                    'title': txn.title,
                    'amount': str(txn.amount),
                    'transaction_type': txn.transaction_type,
                    'date': str(txn.date),
                    'category': txn.category,
                })
    except Exception as e:
        return False, {'error': f'Database error while saving transactions: {str(e)}'}

    return True, {
        'message': f'{len(created)} transactions created successfully.',
        'transactions': created,
    }
